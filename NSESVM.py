from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import operator
import numpy as np
from sklearn.cluster import KMeans
import math
from sklearn.svm import SVR
from sklearn.naive_bayes import MultinomialNB
import random
A=np.zeros((383,495),dtype=np.int)
miRNA_disease=np.loadtxt('knowndiseasemirnainteraction.txt',dtype=np.int)
for pair in miRNA_disease:
    A[pair[1]-1][pair[0]-1]=1
FS=np.loadtxt('miRNA功能类似性矩阵.txt')
FS_weight=np.loadtxt('miRNA功能类似性加权矩阵.txt')
SS1=np.loadtxt('疾病语义类似性矩阵1.txt')
SS2=np.loadtxt('疾病语义类似性矩阵2.txt')
SS_weight=np.loadtxt('疾病语义类似性加权矩阵1.txt')
SS=(SS1+SS2)/2
U_number=[]
for i in range(383):
    for j in range(495):
        if A[i][j]==0:
            U_number.append([j+1,i+1])


def range2rect(x,y,start=0):
    M=[]
    N=[]
    for i in range(x):
        for j in range(y):
            N.append(start)
        M.append(N)
        N=[]
    return M
print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))       
random_number=random.sample(range(0,len(U_number)),5430*2)
U=[]
for i in random_number:
    U.append(U_number[i])
P=[]
for i in range(5430):
    P.append(miRNA_disease[i])
#计算γd
F0=0
for i in range(383):
    for k in range(495):
        F0+=(A[i][k])**2
γ_d=1/(F0/383)
#计算γm
F1=0
for i in range(495):
    for k in range(383):
        F1+=(A[k][i])**2
γ_m=1/(F1/495)

for i in range(383):
    for j in range(383):
        if SS_weight[i][j] == 0:
            F2=(np.linalg.norm(A[i]-A[j]))**2
            KD_d=math.exp(-γ_d*F2)
            SS[i][j]=KD_d
            
#计算miRNA功能相似性权重为零的高斯相似性，并整合。
for i in range(495):
    for j in range(495):
        if FS_weight[i][j] == 0:
            
            F3=(np.linalg.norm(A[:,i]-A[:,j]))**2
            KD_m =math.exp(-γ_m*F3)
            FS[i][j]=KD_m    
#降维开始
Fp=[]
for i in range(383+495):
    Fp.append(0)
Fp=np.array(Fp)
for i in P:
    featurevector=np.append(FS[i[0]-1],SS[i[1]-1])
    featurevector=(featurevector-min(featurevector))/(max(featurevector)-min(featurevector))
    Fp=Fp+featurevector
Fu=[]
for i in range(383+495):
    Fu.append(0)
Fu=np.array(Fu)
for i in U:
    featurevector=np.append(FS[i[0]-1],SS[i[1]-1])
    featurevector=(featurevector-min(featurevector))/(max(featurevector)-min(featurevector))
    Fu=Fu+featurevector   
da=(Fp+Fu)*np.log10(len(P)/Fp+len(U)/Fu)

da_=[]
for i in range(len(da)):
    da_.append([i+1,da[i]])
da_.sort(reverse=True,key=operator.itemgetter(1))
da_495=da_[:495]
da_383=da_[495:]
da_495.sort(reverse=True,key=operator.itemgetter(1))
da_383.sort(reverse=True,key=operator.itemgetter(1))
chose_feature=[]
for i in range(60):
    chose_feature.append(da_495[i][0])
for i in range(40):
    chose_feature.append(da_383[i][0])
chose_feature.sort()
#特征选择
F_V=range2rect(383,495)
for i in range(383):
    for j in range(495):
        vector=np.r_[FS[j],SS[i]]
        max_num=max(np.r_[FS[j],SS[i]])
        min_num=min(np.r_[FS[j],SS[i]])
        vector=(vector-min_num)/(max_num-min_num)
        vector_=[]
        for k in chose_feature:
            vector_.append(vector[k-1])

        F_V[i][j]=np.array(vector_)
Pt=[]
Pr=[]
Pt_vector=[]
random_number=random.sample(range(0,len(P)),int(1*len(P)))
for i in range(len(P)):
    if i in random_number:
        Pt.append(P[i])
        Pt_vector.append(F_V[P[i][1]-1][P[i][0]-1])
    else:
        Pr.append(P[i])
#spy
random_number=random.sample(range(0,len(Pt)),int(0.1*len(Pt)))
spy=[]
Pt_s=[]
for i in range(len(Pt)):
    if i in random_number:
        spy.append(Pt[i])
    else:
        Pt_s.append(Pt[i])
Mix=spy+Pr+U
Mix_lable=[]
X=[]
Y=[]
for i in Pt_s:
    Y.append(1)
    X.append(F_V[i[1]-1][i[0]-1])
for i in Mix:
    Y.append(0)
    X.append(F_V[i[1]-1][i[0]-1])
Mix_copy=X[len(Pt_s):]
clf=MultinomialNB().fit(X[:],Y[:])
revise=clf.predict(Mix_copy)
Mix_copy_lable=[]
for i in revise:
    Mix_copy_lable.append(i)

while Mix_copy_lable!=Mix_lable:
    Mix_lable=Mix_copy_lable[:]
    Y[len(Pt_s):]=Mix_copy_lable[:]
    clf=MultinomialNB().fit(X[:],Y[:])

    revise=clf.predict(Mix_copy)
    Mix_copy_lable=[]
    for i in revise:
        Mix_copy_lable.append(i)

    
#确定阈值t 并根据t选出负样本N
spy_vector=[]
for i in spy:
    spy_vector.append(F_V[i[1]-1][i[0]-1])
identy_spy=clf.predict_proba(spy_vector)
P_s_c1=[]
for i in identy_spy:
    P_s_c1.append(i[1])
P_s_c1.sort()
t=P_s_c1[int(0.15*len(spy))]

N_spy=[]
P_spy=[]
Pr_U_vector=X[len(Pt):]
identy_Pr_U_vector=clf.predict_proba(Pr_U_vector)
for i in range(len(Pr_U_vector)):
    if identy_Pr_U_vector[i][1]<t:
        N_spy.append(Pr_U_vector[i])
    else:
        P_spy.append(Pr_U_vector[i])
#roc
m1=[]
for i in range(len(chose_feature)):
    m1.append(0)
m1=np.array(m1)
m2=[]
for i in range(len(chose_feature)):
    m2.append(0)
m2=np.array(m2)
for i in Pt:
    d=F_V[i[1]-1][i[0]-1]
    m1=m1+d/np.linalg.norm(d)

for i in Pr:
    d=F_V[i[1]-1][i[0]-1]
    m2=m2+d/np.linalg.norm(d)
for i in U:
    d=F_V[i[1]-1][i[0]-1]
    m2=m2+d/np.linalg.norm(d)

c1=(16/len(Pt))*m1-(4/(len(Pr)+len(U)))*m2
c2=(16/(len(Pr)+len(U)))*m2-(4/len(Pt))*m1
N_roc=[]
U_roc=[]
N_roc_lable=[]
for i in Pr:
    d=np.mat(F_V[i[1]-1][i[0]-1])
    sim_c1=d*np.mat(c1).T/((np.linalg.norm(d))*(np.linalg.norm(c1)))
    sim_c2=d*np.mat(c2).T/((np.linalg.norm(d))*(np.linalg.norm(c2)))
    if sim_c1<=sim_c2:
        N_roc.append(i)
    else:
        U_roc.append(i)

for i in U:
    d=np.mat(F_V[i[1]-1][i[0]-1])
    sim_c1=d*np.mat(c1).T/((np.linalg.norm(d))*(np.linalg.norm(c1)))
    sim_c2=d*np.mat(c2).T/((np.linalg.norm(d))*(np.linalg.norm(c2)))
    if sim_c1<=sim_c2:
        N_roc.append(i)
    else:
        U_roc.append(i)
N_two=[]  
U_two=[]  
for i in N_spy:
    d=np.mat(i)
    sim_c1=d*np.mat(c1).T/((np.linalg.norm(d))*(np.linalg.norm(c1)))
    sim_c2=d*np.mat(c2).T/((np.linalg.norm(d))*(np.linalg.norm(c2)))
    if sim_c1<=sim_c2:
        N_two.append(i)
    else:
        U_two.append(i)
P_two=[]
for i in P_spy:
    d=np.mat(i)
    sim_c1=d*np.mat(c1).T/((np.linalg.norm(d))*(np.linalg.norm(c1)))
    sim_c2=d*np.mat(c2).T/((np.linalg.norm(d))*(np.linalg.norm(c2)))
    if sim_c1>sim_c2:
        P_two.append(i)
    else:
        U_two.append(i)

#计算权重: 聚类 计算相似性原型 局部相似性 全局相似性
data = np.array(N_two)
a=round(30*len(N_two)/(len(N_two)+len(U_two)))
estimator = KMeans(n_clusters=a)#构造聚类器
estimator.fit(data)#聚类
label_pred = estimator.labels_ #获取聚类标签
centroids = estimator.cluster_centers_ #获取聚类中心
inertia = estimator.inertia_ # 获取聚类准则的总和
names=locals()
for i in range(1,a+1):
    names['N%s' % i]=[]
for i in range(len(N_two)):
    b= label_pred[i]+1
    names['N%s' % b].append(N_two[i])

def Get_prototypes(N=N1):
    m1=np.zeros(len(chose_feature))
    m2=np.zeros(len(chose_feature))
    for i in N:
        d=np.array(i)
        m2=m2+d/np.linalg.norm(d)
    for i in Pt_vector:
        d=np.array(i)
        m1=m1+d/np.linalg.norm(d)
    for i in P_two:
        d=np.array(i)
        m1=m1+d/np.linalg.norm(d)
    p=(16/(len(Pt_vector)+len(P_two)))*m1-(4/len(N))*m2
    n=(16/len(N))*m2-(4/(len(Pt_vector)+len(P_two)))*m1
    return p,n
for i in range(1,a+1):
    pn=Get_prototypes(N=names['N%s' % i])
    names['p_pro_%s' % i]=pn[0]
    names['n_pro_%s' % i]=pn[1]
data = np.array(U_two)
n=30-a
estimator = KMeans(n_clusters=n)#构造聚类器
estimator.fit(data)#聚类
label_pred = estimator.labels_ #获取聚类标签
centroids = estimator.cluster_centers_ #获取聚类中心
inertia = estimator.inertia_ # 获取聚类准则的总和
names=locals()
for i in range(1,n+1):
    names['U%s' % i]=[]
for i in range(len(U_two)):
    b= label_pred[i]+1
    names['U%s' % b].append(U_two[i])


for i in range(1,n+1):
    tempos=0
    temneg=0
    for x in names['U%s' % i]:
        sim_x_p=[]
        sim_x_n=[]
        for j in range(1,a+1):
            sim_x_n.append(np.dot(x,names['n_pro_%s' % j])/((np.linalg.norm(x))*(np.linalg.norm(names['n_pro_%s' % j]))))
            sim_x_p.append(np.dot(x,names['p_pro_%s' % j])/((np.linalg.norm(x))*(np.linalg.norm(names['p_pro_%s' % j]))))
         
        if max(sim_x_p) > max(sim_x_n):
            tempos+=1
        else:
            temneg+=1
    names['LocP%s' % i]=tempos/len(names['U%s' % i])
    names['LocN%s' % i]=temneg/len(names['U%s' % i])
GloP=[]
GloN=[]
Wp=[]
Wn=[]
for i in range(1,n+1):
    for x in names['U%s' % i]:
        sum_sim_x_p=0
        sum_sim_x_n=0
        for j in range(1,a+1):
            sum_sim_x_n+=(np.dot(x,names['n_pro_%s' % j])/((np.linalg.norm(x))*(np.linalg.norm(names['n_pro_%s' % j]))))
            sum_sim_x_p+=(np.dot(x,names['p_pro_%s' % j])/((np.linalg.norm(x))*(np.linalg.norm(names['p_pro_%s' % j]))))
        GloN.append(sum_sim_x_n/(sum_sim_x_p+sum_sim_x_n))
        Wp.append(0.4*names['LocP%s' % i]+0.6*(sum_sim_x_n/(sum_sim_x_p+sum_sim_x_n)))
        Wn.append(1-0.4*names['LocP%s' % i]-0.6*(sum_sim_x_n/(sum_sim_x_p+sum_sim_x_n)))
Pt_lable=[]
Pt_w=[]
P_two_lable=[]
P_two_w=[]
N_two_lable=[]
N_two_w=[]
U_two_plable=[]
U_two_nlable=[]
for i in range(len(Pt_vector)):
    Pt_lable.append(1)
    Pt_w.append(1)
for i in range(len(P_two)):
    P_two_lable.append(1)
    P_two_w.append(1)
for i in range(len(U_two)):
    U_two_plable.append(1)
    U_two_nlable.append(0)
for i in range(len(N_two)):
    N_two_lable.append(0)
    N_two_w.append(1)

x=Pt_vector+P_two+N_two+U_two+U_two
y=Pt_lable+P_two_lable+N_two_lable+U_two_plable+U_two_nlable
w=Pt_w+P_two_w+N_two_w+Wp+Wn
rf=SVR()
rf.fit(x,y,w)

wb=load_workbook(filename=r'diseasenumber.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
disease_name=[]
for row_A in range(1,384):
    disease_name.append([ws.cell(row=row_A,column=1).value,ws.cell(row=row_A,column=2).value])
wb=load_workbook(filename=r'miRNAnumber.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
miRNA_name=[]
for row_A in range(1,496):
    miRNA_name.append([ws.cell(row=row_A,column=1).value,ws.cell(row=row_A,column=2).value])

print(0)
F_Varray=np.array(F_V)

rds_all=[]
for i in range(383):
    prediction=rf.predict(F_Varray[i,:])
    wb=Workbook()
    ws=wb.active
    ws.title='%s' % disease_name[i][1]
    rds=[]
    for j in range(495):
        if A[i][j]==0:
            rds.append([disease_name[i][1],miRNA_name[j][1],prediction[j]])
    rds.sort(reverse=True,key=operator.itemgetter(2))
    for k in range(len(rds)):
        ws.append(rds[k])
    rds_all+=rds
    #wb.save(filename=r'C:结果\每种疾病的预测结果\%s' % disease_name[i][1] +'.xlsx')
rds_all.sort(reverse=True,key=operator.itemgetter(2))
wb=Workbook()
ws=wb.active
ws.title='所有疾病的预测结果'
for i in range(len(rds_all)):
    ws.append(rds_all[i])
wb.save(filename=r'结果\所有疾病的预测结果.xlsx')
print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))